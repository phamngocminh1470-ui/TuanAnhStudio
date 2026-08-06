"""
Item Bank Manager API — KHKT Research Module
=============================================
FastAPI router phục vụ quản lý ngân hàng câu hỏi đầy đủ metadata.

Endpoints:
  GET    /api/items              - List (với filter + pagination)
  GET    /api/items/stats        - Dashboard statistics
  GET    /api/items/quality-check - Phát hiện câu hỏi lỗi/thiếu metadata
  GET    /api/items/export-csv   - Export CSV
  GET    /api/items/export-xlsx  - Export Excel
  GET    /api/items/{item_id}    - Detail
  POST   /api/items              - Tạo mới
  PUT    /api/items/{item_id}    - Cập nhật
  DELETE /api/items/{item_id}    - Xóa
  POST   /api/items/import-csv   - Import từ CSV
  POST   /api/items/import-xlsx  - Import từ Excel

Dữ liệu lưu trong: backend/irt_item_bank.json (JSON flat-file, không cần DB)
"""

import csv
import io
import json
import os
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator

# ─────────────────────────────────────────────────────────────
# CẤU HÌNH
# ─────────────────────────────────────────────────────────────
ITEM_BANK_PATH = os.path.join(os.path.dirname(__file__), "irt_item_bank.json")

REQUIRED_IMPORT_FIELDS = [
    "question", "option_a", "option_b", "option_c", "option_d",
    "correct_answer", "skill", "topic"
]

VALID_SKILLS = [
    "Tenses", "Passive Voice", "Relative Clauses", "Conditionals",
    "Reported Speech", "Vocabulary", "Collocations", "Pronunciation", "Stress",
    "Word Form & Collocation", "Passive Voice / Word Form", "Locating Specific Information",
    "Reduced Relative Clauses", "Infinitive of Purpose", "Reduced Relative Clauses (Passive)",
    "Word Form", "Collocations / Word Form", "Word Form / Noun Compounds",
    "Vocabulary in Context / Adjectives", "Paragraph Arrangement / Cohesive Devices",
    "Dialogue Arrangement", "Letter/Email Arrangement", "Process Paragraph Arrangement",
    "Speech Structure Arrangement", "Paragraph Arrangement", "Interview Dialogue Arrangement",
    "Argumentative Paragraph Arrangement", "Adverbs / Word Form",
    "Collocations / Commonly Confused Words", "Fixed Adjectives / Comparison",
    "Reflexive Pronouns", "Modal Verbs / Word Form", "Commonly Confused Verbs",
    "Word Form / Adjectives", "Linking Words", "Vocabulary in Context / Parallel Adjectives",
    "Main Idea", "Vocabulary in Context", "Identifying Information NOT Stated",
    "Inference", "Author's Tone & Attitude", "Author's Purpose/Attitude"
]

VALID_TOPICS = [
    "Grammar", "Vocabulary", "Phonology", "Reading", "Writing", "Listening",
    "School Events", "Culture & Community", "Environment & Community",
    "Education & Global Study", "Technology & Education", "School Life",
    "School Life & Communication", "Environment", "Work & Career",
    "Global Education & Languages", "Technology & Society", "Environment & Sustainability",
    "Reading & Lifelong Learning", "Environment & Urban Health", "Culture & Society"
]
VALID_COGNITIVE = ["Remember", "Understand", "Apply", "Analyze", "Evaluate", "Create", "Nhận biết", "Thông hiểu", "Vận dụng"]
VALID_DIFFICULTY = ["Easy", "Medium", "Hard", "Very Hard"]
VALID_STATUS = ["Draft", "Reviewed", "Approved", "Pending Review"]
VALID_CALIBRATION = ["CALIBRATED", "PROVISIONAL", "UNCALIBRATED"]
VALID_QTYPE = [
    "Multiple Choice", "Fill in the blank", "True/False", "Matching",
    "Notice Fill-in", "Leaflet Fill-in", "Sentence Arrangement", "Cloze", "Reading Comprehension"
]

router = APIRouter()


# ─────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def _load_bank() -> Dict[str, Any]:
    """Đọc toàn bộ item bank từ JSON file."""
    if not os.path.exists(ITEM_BANK_PATH):
        return {"schema_version": "2.0", "questions": []}
    with open(ITEM_BANK_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_bank(data: Dict[str, Any]) -> None:
    """Ghi item bank về JSON file (atomic write)."""
    tmp_path = ITEM_BANK_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, ITEM_BANK_PATH)


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _generate_id(existing_ids: List[str]) -> str:
    """Tạo item_id tự động (IRT_Q_XXX) không trùng với existing."""
    nums = []
    for iid in existing_ids:
        m = re.search(r"IRT_Q_(\d+)$", iid)
        if m:
            nums.append(int(m.group(1)))
    next_num = (max(nums) + 1) if nums else 1
    return f"IRT_Q_{next_num:03d}"


def _build_question_dict(
    item_id: str,
    data: Dict[str, Any],
    existing: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Xây dựng câu hỏi đầy đủ metadata từ input.
    existing: dùng để giữ nguyên created_at và các trường không đổi khi update.
    """
    now = _now_iso()
    # options[] backward compat
    options = [
        f"A. {data.get('option_a', '')}",
        f"B. {data.get('option_b', '')}",
        f"C. {data.get('option_c', '')}",
        f"D. {data.get('option_d', '')}",
    ]
    return {
        "item_id": item_id,
        "question": data.get("question", ""),
        "option_a": data.get("option_a", ""),
        "option_b": data.get("option_b", ""),
        "option_c": data.get("option_c", ""),
        "option_d": data.get("option_d", ""),
        "options": options,
        "correct_answer": data.get("correct_answer", "A").upper(),
        "correct": data.get("correct_answer", "A").upper(),  # backward compat
        "explanation": data.get("explanation", ""),
        "topic": data.get("topic", "Grammar"),
        "skill": data.get("skill", "Tenses"),
        "question_type": data.get("question_type", "Multiple Choice"),
        "cognitive_level": data.get("cognitive_level", "Remember"),
        "difficulty_level": data.get("difficulty_level", "Medium"),
        "source": data.get("source", ""),
        "source_year": str(data.get("source_year", "")),
        # IRT parameters — chỉ gán khi được cung cấp và đã hiệu chuẩn
        "calibration_status": data.get("calibration_status", "UNCALIBRATED"),
        "discrimination": data.get("discrimination") if data.get("discrimination") is not None else None,
        "difficulty_parameter": data.get("difficulty_parameter") if data.get("difficulty_parameter") is not None else None,
        "difficulty": data.get("difficulty_parameter") if data.get("difficulty_parameter") is not None else None,  # backward compat
        "guessing_parameter": data.get("guessing_parameter") if data.get("guessing_parameter") is not None else None,
        "guessing": data.get("guessing_parameter") if data.get("guessing_parameter") is not None else None,  # backward compat
        "sample_size": data.get("sample_size", 0),
        # Workflow fields
        "status": data.get("status", "Draft"),
        "reviewer": data.get("reviewer", ""),
        "created_at": existing.get("created_at", now) if existing else now,
        "updated_at": now,
    }


def _apply_filters(
    questions: List[Dict[str, Any]],
    skill: Optional[str],
    topic: Optional[str],
    difficulty_level: Optional[str],
    status: Optional[str],
    calibration_status: Optional[str],
    source: Optional[str],
    q: Optional[str],
) -> List[Dict[str, Any]]:
    result = questions
    if skill:
        result = [x for x in result if x.get("skill") == skill]
    if topic:
        result = [x for x in result if x.get("topic") == topic]
    if difficulty_level:
        result = [x for x in result if x.get("difficulty_level") == difficulty_level]
    if status:
        result = [x for x in result if x.get("status") == status]
    if calibration_status:
        result = [x for x in result if x.get("calibration_status") == calibration_status]
    if source:
        result = [x for x in result if source.lower() in x.get("source", "").lower()]
    if q:
        q_lower = q.lower()
        result = [
            x for x in result
            if q_lower in x.get("question", "").lower()
            or q_lower in x.get("item_id", "").lower()
            or q_lower in x.get("explanation", "").lower()
        ]
    return result


# ─────────────────────────────────────────────────────────────
# PYDANTIC MODELS
# ─────────────────────────────────────────────────────────────

class QuestionCreateRequest(BaseModel):
    question: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str
    explanation: str = ""
    topic: str = "Grammar"
    skill: str = "Tenses"
    question_type: str = "Multiple Choice"
    cognitive_level: str = "Remember"
    difficulty_level: str = "Medium"
    source: str = ""
    source_year: str = ""
    calibration_status: str = "UNCALIBRATED"
    discrimination: Optional[float] = None
    difficulty_parameter: Optional[float] = None
    guessing_parameter: Optional[float] = None
    sample_size: int = 0
    status: str = "Draft"
    reviewer: str = ""

    @field_validator("correct_answer")
    @classmethod
    def validate_correct(cls, v):
        if v.upper() not in ["A", "B", "C", "D"]:
            raise ValueError("correct_answer phải là A, B, C hoặc D")
        return v.upper()

    @field_validator("calibration_status")
    @classmethod
    def validate_calib(cls, v):
        if v not in VALID_CALIBRATION:
            raise ValueError(f"calibration_status phải là: {VALID_CALIBRATION}")
        return v

    @field_validator("status")
    @classmethod
    def validate_status(cls, v):
        if v not in VALID_STATUS:
            raise ValueError(f"status phải là: {VALID_STATUS}")
        return v


class QuestionUpdateRequest(QuestionCreateRequest):
    pass


class BulkStatusUpdate(BaseModel):
    item_ids: List[str]
    status: str
    reviewer: str = ""

    @field_validator("status")
    @classmethod
    def validate_status(cls, v):
        if v not in VALID_STATUS:
            raise ValueError(f"status phải là: {VALID_STATUS}")
        return v


# ─────────────────────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────────────────────

@router.get("/items/stats")
async def get_item_bank_stats():
    """
    Dashboard statistics: tổng số câu, phân bổ theo skill/topic/difficulty/status/calibration.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])

    def _count(field: str) -> Dict[str, int]:
        counts: Dict[str, int] = {}
        for q in questions:
            val = q.get(field, "Unknown") or "Unknown"
            counts[val] = counts.get(val, 0) + 1
        return dict(sorted(counts.items()))

    calibrated_count = sum(1 for q in questions if q.get("calibration_status") == "CALIBRATED")
    approved_count = sum(1 for q in questions if q.get("status") == "Approved")
    missing_explanation = sum(1 for q in questions if not q.get("explanation", "").strip())
    missing_source = sum(1 for q in questions if not q.get("source", "").strip())

    return {
        "status": "success",
        "total": len(questions),
        "calibrated_count": calibrated_count,
        "approved_count": approved_count,
        "missing_explanation": missing_explanation,
        "missing_source": missing_source,
        "by_skill": _count("skill"),
        "by_topic": _count("topic"),
        "by_difficulty_level": _count("difficulty_level"),
        "by_status": _count("status"),
        "by_calibration": _count("calibration_status"),
        "by_source_year": _count("source_year"),
        "by_cognitive_level": _count("cognitive_level"),
    }


@router.get("/items/quality-check")
async def quality_check():
    """
    Kiểm tra chất lượng toàn bộ item bank.
    Trả về danh sách câu có vấn đề cần sửa.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])

    issues: List[Dict[str, Any]] = []
    seen_questions: Dict[str, str] = {}  # question_text -> item_id

    for q in questions:
        q_issues = []
        item_id = q.get("item_id", "?")

        # Kiểm tra các trường bắt buộc
        if not q.get("explanation", "").strip():
            q_issues.append("Thiếu explanation")
        if not q.get("source", "").strip():
            q_issues.append("Thiếu source")
        if not q.get("skill", "").strip():
            q_issues.append("Thiếu skill")
        if not q.get("topic", "").strip():
            q_issues.append("Thiếu topic")
        if not q.get("question", "").strip():
            q_issues.append("Thiếu nội dung câu hỏi")
        if not q.get("option_a") or not q.get("option_b") or not q.get("option_c") or not q.get("option_d"):
            q_issues.append("Thiếu đáp án (option_a/b/c/d)")

        # Kiểm tra calibration
        if q.get("calibration_status") == "UNCALIBRATED":
            q_issues.append("Chưa hiệu chuẩn IRT (UNCALIBRATED)")
        if q.get("calibration_status") == "PROVISIONAL":
            q_issues.append("Tham số IRT tạm thời (PROVISIONAL) — cần hiệu chuẩn")
        if q.get("calibration_status") == "CALIBRATED" and q.get("difficulty_parameter") is None:
            q_issues.append("CALIBRATED nhưng thiếu difficulty_parameter")

        # Kiểm tra workflow status
        if q.get("status") == "Draft":
            q_issues.append("Chưa được kiểm duyệt (Draft)")

        # Kiểm tra trùng lặp
        q_text = q.get("question", "").strip().lower()
        if q_text:
            if q_text in seen_questions:
                q_issues.append(f"Trùng câu hỏi với {seen_questions[q_text]}")
            else:
                seen_questions[q_text] = item_id

        if q_issues:
            issues.append({
                "item_id": item_id,
                "question_preview": (q.get("question", "")[:80] + "...") if len(q.get("question", "")) > 80 else q.get("question", ""),
                "skill": q.get("skill", ""),
                "status": q.get("status", ""),
                "calibration_status": q.get("calibration_status", ""),
                "issues": q_issues,
                "issue_count": len(q_issues)
            })

    # Sắp xếp theo số lỗi giảm dần
    issues.sort(key=lambda x: x["issue_count"], reverse=True)

    return {
        "status": "success",
        "total_questions": len(questions),
        "questions_with_issues": len(issues),
        "clean_questions": len(questions) - len(issues),
        "issues": issues
    }


@router.get("/items")
async def list_items(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    skill: Optional[str] = None,
    topic: Optional[str] = None,
    difficulty_level: Optional[str] = None,
    status: Optional[str] = None,
    calibration_status: Optional[str] = None,
    source: Optional[str] = None,
    q: Optional[str] = None,  # search query
):
    """
    Lấy danh sách câu hỏi với filter, search và pagination.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])

    filtered = _apply_filters(questions, skill, topic, difficulty_level, status, calibration_status, source, q)

    total = len(filtered)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    end = start + page_size
    page_items = filtered[start:end]

    return {
        "status": "success",
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "items": page_items
    }


@router.get("/items/export-csv")
async def export_csv(
    skill: Optional[str] = None,
    topic: Optional[str] = None,
    status: Optional[str] = None,
):
    """
    Export câu hỏi ra file CSV (UTF-8 with BOM cho Excel đọc được).
    """
    bank = _load_bank()
    questions = bank.get("questions", [])
    filtered = _apply_filters(questions, skill, topic, None, status, None, None, None)

    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "item_id", "question", "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "explanation", "topic", "skill", "question_type",
        "cognitive_level", "difficulty_level", "source", "source_year",
        "calibration_status", "discrimination", "difficulty_parameter", "guessing_parameter",
        "sample_size", "status", "reviewer", "created_at", "updated_at"
    ]
    writer.writerow(headers)

    for q in filtered:
        writer.writerow([
            q.get("item_id", ""),
            q.get("question", ""),
            q.get("option_a", ""),
            q.get("option_b", ""),
            q.get("option_c", ""),
            q.get("option_d", ""),
            q.get("correct_answer", ""),
            q.get("explanation", ""),
            q.get("topic", ""),
            q.get("skill", ""),
            q.get("question_type", "Multiple Choice"),
            q.get("cognitive_level", ""),
            q.get("difficulty_level", ""),
            q.get("source", ""),
            q.get("source_year", ""),
            q.get("calibration_status", "UNCALIBRATED"),
            q.get("discrimination", ""),
            q.get("difficulty_parameter", ""),
            q.get("guessing_parameter", ""),
            q.get("sample_size", 0),
            q.get("status", "Draft"),
            q.get("reviewer", ""),
            q.get("created_at", ""),
            q.get("updated_at", ""),
        ])

    csv_content = "\ufeff" + output.getvalue()  # UTF-8 BOM for Excel

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"item_bank_export_{timestamp}.csv"

    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/items/export-xlsx")
async def export_xlsx(
    skill: Optional[str] = None,
    topic: Optional[str] = None,
    status: Optional[str] = None,
):
    """
    Export câu hỏi ra file Excel (.xlsx).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Thư viện openpyxl chưa được cài đặt. Chạy: pip install openpyxl")

    bank = _load_bank()
    questions = bank.get("questions", [])
    filtered = _apply_filters(questions, skill, topic, None, status, None, None, None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Bank"

    # Header style
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        ("item_id", 14), ("question", 60), ("option_a", 25), ("option_b", 25),
        ("option_c", 25), ("option_d", 25), ("correct_answer", 14),
        ("explanation", 50), ("topic", 16), ("skill", 20), ("question_type", 18),
        ("cognitive_level", 18), ("difficulty_level", 16), ("source", 30),
        ("source_year", 12), ("calibration_status", 18), ("discrimination", 15),
        ("difficulty_parameter", 20), ("guessing_parameter", 18), ("sample_size", 13),
        ("status", 12), ("reviewer", 20), ("created_at", 22), ("updated_at", 22)
    ]

    for col_idx, (header_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_width

    ws.row_dimensions[1].height = 22

    # Data rows
    STATUS_COLORS = {
        "Approved": "D4EDDA",
        "Reviewed": "FFF3CD",
        "Draft": "F8D7DA"
    }

    for row_idx, q in enumerate(filtered, 2):
        row_data = [
            q.get("item_id", ""), q.get("question", ""),
            q.get("option_a", ""), q.get("option_b", ""),
            q.get("option_c", ""), q.get("option_d", ""),
            q.get("correct_answer", ""), q.get("explanation", ""),
            q.get("topic", ""), q.get("skill", ""),
            q.get("question_type", "Multiple Choice"), q.get("cognitive_level", ""),
            q.get("difficulty_level", ""), q.get("source", ""),
            q.get("source_year", ""), q.get("calibration_status", "UNCALIBRATED"),
            q.get("discrimination", ""), q.get("difficulty_parameter", ""),
            q.get("guessing_parameter", ""), q.get("sample_size", 0),
            q.get("status", "Draft"), q.get("reviewer", ""),
            q.get("created_at", ""), q.get("updated_at", "")
        ]

        q_status = q.get("status", "Draft")
        row_color = STATUS_COLORS.get(q_status, "FFFFFF")

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

        ws.row_dimensions[row_idx].height = 50

    # Freeze header row
    ws.freeze_panes = "A2"

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"item_bank_export_{timestamp}.xlsx"

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/items/export-template-csv")
async def export_csv_template():
    """
    Download file CSV mẫu để giáo viên điền câu hỏi mới và import lại.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "question", "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "explanation", "topic", "skill",
        "question_type", "cognitive_level", "difficulty_level",
        "source", "source_year", "status", "reviewer"
    ]
    writer.writerow(headers)

    # 1 dòng ví dụ
    writer.writerow([
        "She usually ________ to school by bicycle every morning.",
        "goes", "go", "is going", "went",
        "A",
        "Dùng thì Hiện tại đơn với 'usually'. Chủ ngữ 'She' thêm '-s': goes.",
        "Grammar", "Tenses",
        "Multiple Choice", "Remember", "Easy",
        "Đề thi tốt nghiệp THPT", "2023", "Draft", ""
    ])

    csv_content = "\ufeff" + output.getvalue()

    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="item_bank_template.csv"'}
    )


@router.get("/items/{item_id}")
async def get_item(item_id: str):
    """
    Lấy chi tiết một câu hỏi theo item_id.
    """
    bank = _load_bank()
    for q in bank.get("questions", []):
        if q.get("item_id") == item_id:
            return {"status": "success", "item": q}
    raise HTTPException(status_code=404, detail=f"Không tìm thấy câu hỏi '{item_id}'")


@router.post("/items")
async def create_item(request: QuestionCreateRequest):
    """
    Tạo câu hỏi mới. item_id được tự động tạo.
    Mặc định status=Draft, calibration_status=UNCALIBRATED.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])
    existing_ids = [q.get("item_id", "") for q in questions]

    new_id = _generate_id(existing_ids)
    new_q = _build_question_dict(new_id, request.model_dump())

    questions.append(new_q)
    bank["questions"] = questions
    _save_bank(bank)

    return {
        "status": "success",
        "message": f"Đã tạo câu hỏi '{new_id}' thành công.",
        "item": new_q
    }


@router.put("/items/{item_id}")
async def update_item(item_id: str, request: QuestionUpdateRequest):
    """
    Cập nhật câu hỏi theo item_id.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])

    for idx, q in enumerate(questions):
        if q.get("item_id") == item_id:
            updated = _build_question_dict(item_id, request.model_dump(), existing=q)
            questions[idx] = updated
            bank["questions"] = questions
            _save_bank(bank)
            return {
                "status": "success",
                "message": f"Đã cập nhật câu hỏi '{item_id}'.",
                "item": updated
            }

    raise HTTPException(status_code=404, detail=f"Không tìm thấy câu hỏi '{item_id}'")


@router.delete("/items/{item_id}")
async def delete_item(item_id: str):
    """
    Xóa câu hỏi theo item_id. Thao tác không thể hoàn tác.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])
    original_count = len(questions)
    questions = [q for q in questions if q.get("item_id") != item_id]

    if len(questions) == original_count:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy câu hỏi '{item_id}'")

    bank["questions"] = questions
    _save_bank(bank)
    return {"status": "success", "message": f"Đã xóa câu hỏi '{item_id}'."}


@router.post("/items/bulk/status")
async def bulk_update_status(request: BulkStatusUpdate):
    """
    Cập nhật trạng thái (Draft/Reviewed/Approved) cho nhiều câu hỏi cùng lúc.
    """
    bank = _load_bank()
    questions = bank.get("questions", [])
    updated_ids = []
    now = _now_iso()

    for q in questions:
        if q.get("item_id") in request.item_ids:
            q["status"] = request.status
            if request.reviewer:
                q["reviewer"] = request.reviewer
            q["updated_at"] = now
            updated_ids.append(q["item_id"])

    bank["questions"] = questions
    _save_bank(bank)

    return {
        "status": "success",
        "updated_count": len(updated_ids),
        "updated_ids": updated_ids,
        "new_status": request.status
    }


@router.post("/items/import-csv")
async def import_csv(file: UploadFile = File(...)):
    """
    Import câu hỏi từ file CSV.
    Kiểm tra validation nghiêm ngặt, báo lỗi từng dòng.
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="File phải có đuôi .csv")

    content = await file.read()
    # Thử decode UTF-8 (có hoặc không có BOM)
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1258")
        except Exception:
            raise HTTPException(status_code=400, detail="Không thể đọc file CSV. Vui lòng lưu dưới dạng UTF-8.")

    reader = csv.DictReader(io.StringIO(text))

    errors = []
    valid_rows = []

    for line_num, row in enumerate(reader, start=2):
        row_errors = []

        # Kiểm tra trường bắt buộc
        for field in REQUIRED_IMPORT_FIELDS:
            if not row.get(field, "").strip():
                row_errors.append(f"Thiếu trường '{field}'")

        # Kiểm tra correct_answer
        correct = row.get("correct_answer", "").strip().upper()
        if correct not in ["A", "B", "C", "D"]:
            row_errors.append(f"correct_answer='{correct}' không hợp lệ (phải là A/B/C/D)")

        if row_errors:
            errors.append({"row": line_num, "errors": row_errors})
        else:
            valid_rows.append(row)

    if errors:
        return {
            "status": "error",
            "message": f"Có {len(errors)} dòng lỗi. Vui lòng sửa trước khi import.",
            "errors": errors,
            "valid_count": len(valid_rows)
        }

    # Import các dòng hợp lệ
    bank = _load_bank()
    questions = bank.get("questions", [])
    existing_ids = [q.get("item_id", "") for q in questions]

    imported = []
    for row in valid_rows:
        # Tìm IDs dư thừa từ hàng trước
        all_ids = existing_ids + [q["item_id"] for q in imported]
        new_id = _generate_id(all_ids)

        new_q = _build_question_dict(new_id, {
            "question": row.get("question", "").strip(),
            "option_a": row.get("option_a", "").strip(),
            "option_b": row.get("option_b", "").strip(),
            "option_c": row.get("option_c", "").strip(),
            "option_d": row.get("option_d", "").strip(),
            "correct_answer": row.get("correct_answer", "A").strip().upper(),
            "explanation": row.get("explanation", "").strip(),
            "topic": row.get("topic", "Grammar").strip(),
            "skill": row.get("skill", "Tenses").strip(),
            "question_type": row.get("question_type", "Multiple Choice").strip() or "Multiple Choice",
            "cognitive_level": row.get("cognitive_level", "Remember").strip() or "Remember",
            "difficulty_level": row.get("difficulty_level", "Medium").strip() or "Medium",
            "source": row.get("source", "").strip(),
            "source_year": row.get("source_year", "").strip(),
            "calibration_status": "UNCALIBRATED",  # Import luôn UNCALIBRATED
            "status": row.get("status", "Draft").strip() or "Draft",
            "reviewer": row.get("reviewer", "").strip(),
        })
        imported.append(new_q)

    questions.extend(imported)
    bank["questions"] = questions
    _save_bank(bank)

    return {
        "status": "success",
        "message": f"Import thành công {len(imported)} câu hỏi.",
        "imported_count": len(imported),
        "imported_ids": [q["item_id"] for q in imported]
    }


@router.post("/items/import-xlsx")
async def import_xlsx(file: UploadFile = File(...)):
    """
    Import câu hỏi từ file Excel (.xlsx).
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File phải có đuôi .xlsx")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Thư viện openpyxl chưa được cài đặt. Chạy: pip install openpyxl")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể đọc file Excel: {str(e)}")

    # Đọc header từ dòng 1
    headers = [cell.value for cell in ws[1]]
    if not headers or not any(h for h in headers):
        raise HTTPException(status_code=400, detail="File Excel không có header dòng 1.")

    headers = [str(h).strip() if h else "" for h in headers]

    errors = []
    valid_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # Bỏ qua dòng trống

        row_dict = {headers[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else "" for i in range(len(headers))}
        row_errors = []

        for field in REQUIRED_IMPORT_FIELDS:
            if not row_dict.get(field, "").strip():
                row_errors.append(f"Thiếu trường '{field}'")

        correct = row_dict.get("correct_answer", "").strip().upper()
        if correct not in ["A", "B", "C", "D"]:
            row_errors.append(f"correct_answer='{correct}' không hợp lệ")

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})
        else:
            valid_rows.append(row_dict)

    if errors:
        return {
            "status": "error",
            "message": f"Có {len(errors)} dòng lỗi. Vui lòng sửa trước khi import.",
            "errors": errors,
            "valid_count": len(valid_rows)
        }

    bank = _load_bank()
    questions = bank.get("questions", [])
    existing_ids = [q.get("item_id", "") for q in questions]

    imported = []
    for row in valid_rows:
        all_ids = existing_ids + [q["item_id"] for q in imported]
        new_id = _generate_id(all_ids)
        new_q = _build_question_dict(new_id, {**row, "calibration_status": "UNCALIBRATED"})
        imported.append(new_q)

    questions.extend(imported)
    bank["questions"] = questions
    _save_bank(bank)

    return {
        "status": "success",
        "message": f"Import thành công {len(imported)} câu hỏi từ Excel.",
        "imported_count": len(imported),
        "imported_ids": [q["item_id"] for q in imported]
    }
