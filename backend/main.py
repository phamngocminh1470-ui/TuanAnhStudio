import sys
import os
import json

# BUG-11 FIX: Xóa import trùng lặp io và os
# Đảm bảo in ra màn hình console Windows không bị lỗi font Unicode
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from sqlmodel import Session, select
from collections import defaultdict
import io

from ai_services import (
    chat_with_gemini, text_to_speech, speech_to_text, assess_pronunciation, 
    generate_adaptive_question_with_gemini, generate_pronunciation_sentence_with_gemini,
    generate_adaptive_reading, generate_adaptive_listening, calculate_predicted_exam_scores, evaluate_writing_with_gemini,
    generate_writing_sample_with_gemini, solve_exam_by_image, execute_writing_ai_prompt, clean_api_key
)
from adaptive_learning import (
    IRTEngine, IRTQuestion, SpacedRepetitionEngine, ItemBank, KnowledgeGraph, 
    AdaptiveQuestionSelector, DiagnosticEngine, SM2Engine, FSRSEngine, ResearchLogger,
    compute_skill_mastery, LearningPathEngine
)
from item_bank_api import router as item_bank_router
from auth_api import router as auth_router
from user_progress_api import router as user_progress_router
from content_api import router as content_router
from sqlmodel import Session, select
from database import create_db_and_tables, get_session, User, engine, VocabularyWord

app = FastAPI(
    title="AI English Mentor API",
    description="Backend API phục vụ hệ thống Gia sư AI tiếng Anh cá nhân hóa thích ứng",
    version="2.0.0"
)

# Khởi tạo SQLite database khi app start
@app.on_event("startup")
def on_startup():
    create_db_and_tables()

# Cấu hình CORS để Frontend (React/Vite) kết nối được
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Trong thực tế nên giới hạn địa chỉ frontend
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include routers
app.include_router(auth_router, prefix="/api", tags=["Authentication"])
app.include_router(user_progress_router, prefix="/api", tags=["User Progress"])
app.include_router(item_bank_router, prefix="/api", tags=["Item Bank Manager"])
app.include_router(content_router, prefix="/api")


# ----------------- ĐỊNH NGHĨA DỮ LIỆU ĐẦU VÀO -----------------
class Message(BaseModel):
    role: str  # "user" hoặc "model" (theo chuẩn Gemini)
    content: str

class ChatRequest(BaseModel):
    messages: List[Message]
    system_instruction: Optional[str] = (
        "You are an encouraging and patient AI English Mentor named Antigravity. "
        "Your task is to help Vietnamese students from grade 6 to 12 learn English. "
        "Always respond in clear, grammatically correct English. "
        "If the student speaks Vietnamese, you can guide them back to English gently. "
        "Keep your responses concise, friendly, and pedagogically sound (explain grammar/vocabulary if they make mistakes)."
    )

class TTSRequest(BaseModel):
    text: str

class WritingPracticeRequest(BaseModel):
    prompt: str



# ----------------- ĐỊNH NGHĨA ENDPOINTS -----------------

@app.get("/")
async def root():
    """Trang chủ API Backend"""
    return {
        "message": "AI English Mentor Backend API is running successfully!",
        "docs": "/docs",
        "health": "/api/health"
    }


@app.get("/api/health")
async def health_check():
    """Kiểm tra trạng thái máy chủ"""
    return {
        "status": "healthy",
        "gemini_configured": bool(os.getenv("GEMINI_API_KEY")),
        "groq_configured": bool(os.getenv("GROQ_API_KEY")),
        "azure_speech_configured": bool(os.getenv("AZURE_SPEECH_KEY"))
    }


class SaveKeysRequest(BaseModel):
    gemini: Optional[str] = ""
    groq: Optional[str] = ""
    azure: Optional[str] = ""

@app.post("/api/save-keys")
async def save_keys_endpoint(request: SaveKeysRequest):
    """
    Lưu khóa API vào tệp .env trên máy chủ và cập nhật runtime memory cho toàn hệ thống.
    """
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    
    if request.gemini and request.gemini.strip():
        os.environ["GEMINI_API_KEY"] = request.gemini.strip()
    if request.groq and request.groq.strip():
        os.environ["GROQ_API_KEY"] = request.groq.strip()
    if request.azure and request.azure.strip():
        os.environ["AZURE_SPEECH_KEY"] = request.azure.strip()
        
    try:
        lines = [
            f"GEMINI_API_KEY={os.environ.get('GEMINI_API_KEY', '')}\n",
            f"GROQ_API_KEY={os.environ.get('GROQ_API_KEY', '')}\n",
            f"AZURE_SPEECH_KEY={os.environ.get('AZURE_SPEECH_KEY', '')}\n",
            f"AZURE_SPEECH_REGION={os.environ.get('AZURE_SPEECH_REGION', 'southeastasia')}\n"
        ]
        with open(env_path, "w", encoding="utf-8") as f:
            f.writelines(lines)
            
        return {"status": "success", "message": "Đã lưu API Keys lên máy chủ thành công!"}
    except Exception as e:
        return {"status": "partial", "message": f"Đã nạp vào bộ nhớ (lỗi ghi file: {e})"}



@app.post("/api/chat")
async def chat_endpoint(
    request: ChatRequest,
    x_gemini_key: Optional[str] = Header(None),
    x_groq_key: Optional[str] = Header(None)
):
    """
    Endpoint nhận cuộc hội thoại và trả về phản hồi từ Socrates AI (Gemini + Groq + Socratic Knowledge Engine).
    """
    messages_dict = [{"role": msg.role, "content": msg.content} for msg in request.messages]
    
    try:
        reply = await chat_with_gemini(
            messages_dict,
            system_instruction=request.system_instruction,
            custom_key=x_gemini_key,
            custom_groq_key=x_groq_key
        )
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi xử lý chatbot: {str(e)}")


@app.post("/api/writing/practice-ai")
async def writing_practice_ai_endpoint(
    request: WritingPracticeRequest,
    x_gemini_key: Optional[str] = Header(None),
    x_groq_key: Optional[str] = Header(None)
):
    """
    Proxy endpoint gọi AI cho Luyện viết câu, Chấm luận 4 tiêu chí, Dàn ý & Bài mẫu.
    """
    try:
        reply = await execute_writing_ai_prompt(
            prompt=request.prompt,
            custom_gemini_key=x_gemini_key,
            custom_groq_key=x_groq_key
        )
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tts")
async def tts_endpoint(request: TTSRequest, x_azure_key: Optional[str] = Header(None)):
    """
    Endpoint chuyển đổi text sang giọng nói và trả về stream file âm thanh MP3.
    """
    if not request.text.strip():
        raise HTTPException(status_code=400, detail="Văn bản không được để trống")
        
    try:
        audio_data = await text_to_speech(request.text, custom_key=x_azure_key)
        return StreamingResponse(io.BytesIO(audio_data), media_type="audio/mpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý TTS: {str(e)}")


@app.post("/api/stt")
async def stt_endpoint(file: UploadFile = File(...), x_groq_key: Optional[str] = Header(None)):
    """
    Endpoint nhận file ghi âm từ frontend và chuyển đổi thành văn bản bằng Whisper.
    """
    try:
        # Đọc dữ liệu file âm thanh gửi lên
        audio_bytes = await file.read()
        transcription = await speech_to_text(audio_bytes, filename=file.filename, custom_key=x_groq_key)
        return {"text": transcription}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý Speech-to-Text: {str(e)}")


@app.post("/api/pronounce-assess")
async def pronounce_assess_endpoint(
    reference_text: str = Form(...),
    file: UploadFile = File(...),
    x_azure_key: Optional[str] = Header(None),
    x_gemini_key: Optional[str] = Header(None)
):
    """
    Endpoint nhận file ghi âm và văn bản chuẩn để đánh giá phát âm chi tiết.
    """
    try:
        audio_bytes = await file.read()
        assessment_result = await assess_pronunciation(
            audio_bytes, 
            reference_text, 
            custom_key=x_azure_key, 
            custom_gemini_key=x_gemini_key
        )
        return assessment_result
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý chấm phát âm: {str(e)}")



# Pydantic models cho thuật toán thích ứng (Adaptive Learning)
class IRTQuestionModel(BaseModel):
    item_id: str
    difficulty: float
    discrimination: float = 1.0
    guessing: float = 0.2

class NextQuestionRequest(BaseModel):
    theta: float
    excluded_ids: List[str]
    pool: List[IRTQuestionModel]

class IRTHistoryItem(BaseModel):
    question: IRTQuestionModel
    response: int # 1 = Đúng, 0 = Sai

class UpdateAbilityRequest(BaseModel):
    history: List[IRTHistoryItem]
    student_id: Optional[str] = "anonymous"
    experiment_group: Optional[str] = "ADAPTIVE"
    repetition_engine: Optional[str] = "SM2"

class SpacedRepetitionRequest(BaseModel):
    quality: int
    current_repetition: int
    current_ef: float
    current_interval: int
    stability: Optional[float] = 2.0
    difficulty: Optional[float] = 3.0
    days_since_last: Optional[int] = 1
    engine: str = "SM2"

@app.post("/api/adaptive/next-question")
async def get_next_question(request: NextQuestionRequest):
    """
    API chọn câu hỏi thích ứng tiếp theo dựa trên năng lực hiện tại của người học (Fisher Information).
    """
    try:
        pool_objects = [
            IRTQuestion(
                item_id=q.item_id,
                difficulty=q.difficulty,
                discrimination=q.discrimination,
                guessing=q.guessing
            ) for q in request.pool
        ]
        
        next_q = IRTEngine.select_next_question(
            theta=request.theta,
            pool=pool_objects,
            excluded_ids=request.excluded_ids
        )
        
        if next_q:
            return {
                "status": "success",
                "question": {
                    "item_id": next_q.item_id,
                    "difficulty": next_q.b,
                    "discrimination": next_q.a,
                    "guessing": next_q.c
                }
            }
        else:
            return {"status": "finished", "message": "Đã hoàn thành toàn bộ câu hỏi trong kho."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi chọn câu hỏi thích ứng: {str(e)}")

@app.post("/api/adaptive/update-ability")
async def update_ability(request: UpdateAbilityRequest):
    """
    API tính toán lại năng lực (theta) của học sinh dựa trên EAP.
    BUG-01 FIX: Dùng hàm compute_skill_mastery() dùng chung, chỉ tính từ item bank thực tế.
    BUG-03 FIX: theta_before được tính từ history[:-1] (trước câu cuối), không phải new_theta ± 0.15.
    """
    try:
        bank = ItemBank()
        history_objects = []

        for item in request.history:
            q_id = item.question.item_id
            if q_id in bank.questions:
                q_ref = bank.questions[q_id]
                q_obj = IRTQuestion(
                    item_id=q_id,
                    difficulty=q_ref.b,
                    discrimination=q_ref.a,
                    guessing=q_ref.c
                )
            else:
                q_obj = IRTQuestion(
                    item_id=q_id,
                    difficulty=item.question.difficulty,
                    discrimination=item.question.discrimination,
                    guessing=item.question.guessing
                )
            history_objects.append((q_obj, item.response))

        # BUG-03 FIX: Tính theta_before từ history không bao gồm câu cuối
        theta_before = IRTEngine.estimate_ability_eap(history_objects[:-1]) if len(history_objects) > 1 else 0.0
        new_theta = IRTEngine.estimate_ability_eap(history_objects)

        # BUG-01 FIX: Dùng hàm compute_skill_mastery() được tách ra, tránh fallback sai kỹ năng
        history_for_mastery = [
            {"itemId": item.question.item_id, "result": item.response}
            for item in request.history
        ]
        skill_mastery = compute_skill_mastery(history_for_mastery, bank)

        # Chẩn đoán lỗi nếu câu cuối cùng bị sai
        diagnostics = None
        if request.history:
            last_item = request.history[-1]
            if last_item.response == 0:
                q_id = last_item.question.item_id
                if q_id in bank.questions:
                    diagnostics = DiagnosticEngine.diagnose_error(
                        bank.questions[q_id], "wrong", skill_mastery
                    )

        # Ghi log nghiên cứu thực nghiệm
        if request.history:
            last_item = request.history[-1]
            q_id = last_item.question.item_id
            last_q_skill = bank.questions[q_id].skill if q_id in bank.questions else "Unknown"
            # Tính mastery trước câu cuối để log chính xác
            history_before_last = history_for_mastery[:-1]
            mastery_before = compute_skill_mastery(history_before_last, bank)

            logger = ResearchLogger()
            logger.log_session({
                "student_id": request.student_id,
                "experiment_group": request.experiment_group,
                "repetition_engine": request.repetition_engine,
                "question_id": q_id,
                "skill": last_q_skill,
                "correct": last_item.response,
                "theta_before": theta_before,
                "theta_after": new_theta,
                "mastery_before": mastery_before,
                "mastery_after": skill_mastery,
                "recommendation_reason": "IRT Adaptive CAT Session"
            })

        return {
            "status": "success",
            "new_theta": new_theta,
            "skill_mastery": skill_mastery,
            "diagnostics": diagnostics
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi ước lượng năng lực học sinh: {str(e)}")

@app.post("/api/spaced-repetition/next-review")
async def get_next_review(request: SpacedRepetitionRequest):
    """
    API tính toán lịch ôn tập tiếp theo hỗ trợ chuyển đổi linh hoạt giữa SM-2 và FSRS phục vụ nghiên cứu thực nghiệm.
    """
    try:
        if request.engine == "FSRS":
            stability = request.stability if request.stability else 2.0
            difficulty = request.difficulty if request.difficulty else 3.0
            days = request.days_since_last if request.days_since_last else 1
            
            interval, new_stability, new_difficulty = FSRSEngine.calculate(
                quality=request.quality,
                stability=stability,
                difficulty=difficulty,
                days_since_last=days
            )
            return {
                "status": "success",
                "engine": "FSRS",
                "next_interval_days": interval,
                "new_stability": new_stability,
                "new_difficulty": new_difficulty
            }
        else:
            interval, ef, repetition = SM2Engine.calculate(
                quality=request.quality,
                repetition=request.current_repetition,
                ef=request.current_ef,
                interval=request.current_interval
            )
            return {
                "status": "success",
                "engine": "SM2",
                "next_interval_days": interval,
                "new_ef": ef,
                "new_repetition": repetition
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi tính toán lặp lại ngắt quãng: {str(e)}")


class GenerateQuestionRequest(BaseModel):
    grade: str = "10"
    theta: float = 0.0
    history: Optional[List[Dict[str, Any]]] = []
    topic: Optional[str] = None
    part: Optional[str] = None
    difficulty: Optional[str] = None
    exclude_id: Optional[str] = None

@app.post("/api/adaptive/generate-question")
async def generate_question(request: GenerateQuestionRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    Lựa chọn câu hỏi thích ứng tiếp theo từ Item Bank.
    Hỗ trợ lọc theo Chủ đề, Dạng bài (Part 1 - 3), Mức độ khó và đổi câu hỏi tức thì.
    """
    try:
        bank = ItemBank()
        graph = KnowledgeGraph()
        selector = AdaptiveQuestionSelector(bank, graph)

        history_items = request.history if request.history else []
        skill_mastery = compute_skill_mastery(history_items, bank)
        formatted_history = [
            {"itemId": h.get("itemId", ""), "result": h.get("result", 0)}
            for h in history_items
        ]

        # Chọn câu tiếp theo bằng selector với đầy đủ bộ lọc
        next_q, reason = selector.select_question(
            theta=request.theta,
            irt_history=formatted_history,
            skill_mastery=skill_mastery,
            topic_filter=request.topic,
            part_filter=request.part,
            difficulty_filter=request.difficulty,
            exclude_id=request.exclude_id
        )

        if next_q:
            explanation = next_q.explanation if next_q.explanation else f"Đáp án đúng là {next_q.correct}."
            return {
                "status": "success",
                "question": {
                    "item_id": next_q.item_id,
                    "task_type": next_q.task_type,
                    "question": next_q.question_text,
                    "options": next_q.options,
                    "correct": next_q.correct,
                    "statements": next_q.statements,
                    "correct_short": next_q.correct_short,
                    "passage": next_q.passage,
                    "passage_id": next_q.passage_id,
                    "difficulty": next_q.b,
                    "discrimination": next_q.a,
                    "guessing": next_q.c,
                    "skill": next_q.skill,
                    "topic": next_q.topic,
                    "explanation": explanation
                },
                "recommendation_reason": reason,
                "skill_mastery": skill_mastery
            }
        else:
            return {
                "status": "completed",
                "message": "Đã hoàn thành đánh giá toàn bộ câu hỏi định chuẩn trong Ngân hàng câu hỏi (Item Bank).",
                "question": None,
                "recommendation_reason": "Ngân hàng câu hỏi định chuẩn đã được kiểm duyệt hoàn toàn. Đánh giá năng lực đạt độ hội tụ.",
                "skill_mastery": skill_mastery
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi chọn câu hỏi thích ứng: {str(e)}")


class LearningPathRequest(BaseModel):
    theta: float = 0.0
    skill_mastery: Optional[Dict[str, float]] = None
    history: Optional[List[Dict[str, Any]]] = []


@app.post("/api/adaptive/learning-path")
async def get_learning_path(request: LearningPathRequest):
    """
    BUG-12: Learning Path Engine hoàn toàn deterministic.
    Trả về kế hoạch học cá nhân hóa dựa trên theta, skill mastery, knowledge graph và history.
    Không sử dụng Gemini hay LLM để ra quyết định.
    """
    try:
        bank = ItemBank()
        graph = KnowledgeGraph()

        # Dùng skill_mastery từ request nếu có, nếu không tính từ history
        if request.skill_mastery:
            skill_mastery = request.skill_mastery
        else:
            history_items = request.history if request.history else []
            skill_mastery = compute_skill_mastery(history_items, bank)

        daily_plan = LearningPathEngine.generate_daily_plan(
            theta=request.theta,
            skill_mastery=skill_mastery,
            knowledge_graph=graph,
            item_bank=bank,
            history=request.history if request.history else []
        )

        return {
            "status": "success",
            "theta": request.theta,
            "skill_mastery": skill_mastery,
            "daily_plan": daily_plan,
            "algorithm": "Deterministic LearningPathEngine v1.0 (IRT + Mastery + KnowledgeGraph)"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi sinh lộ trình học: {str(e)}")


@app.get("/api/teacher/report")
async def get_teacher_report():
    """
    BUG-13: Teacher View — Báo cáo tổng hợp cho giáo viên.
    Đọc research_experiment_logs.jsonl và trả về thống kê tổng hợp.
    Không yêu cầu xác thực (demo mode).
    """
    try:
        logger = ResearchLogger()
        logs = logger.read_all_logs()

        if not logs:
            return {
                "status": "success",
                "total_sessions": 0,
                "message": "Chưa có dữ liệu học tập. Hãy mời học sinh sử dụng hệ thống trước.",
                "students": [],
                "skill_accuracy": {}
            }

        # Tổng hợp theo student_id
        student_data: Dict[str, Any] = {}
        skill_stats: Dict[str, Dict[str, int]] = {}

        for log in logs:
            sid = log.get("student_id", "anonymous")
            skill = log.get("skill", "Unknown")
            correct = int(log.get("correct", 0))
            theta_after = float(log.get("theta_after", 0.0))

            # Theo dõi từng học sinh
            if sid not in student_data:
                student_data[sid] = {
                    "student_id": sid,
                    "experiment_group": log.get("experiment_group", "ADAPTIVE"),
                    "total_questions": 0,
                    "correct_count": 0,
                    "latest_theta": theta_after,
                    "sessions": 0
                }
            student_data[sid]["total_questions"] += 1
            student_data[sid]["correct_count"] += correct
            student_data[sid]["latest_theta"] = theta_after
            student_data[sid]["sessions"] = student_data[sid]["total_questions"]

            # Thống kê theo kỹ năng
            if skill not in skill_stats:
                skill_stats[skill] = {"total": 0, "correct": 0}
            skill_stats[skill]["total"] += 1
            skill_stats[skill]["correct"] += correct

        # Tính tỷ lệ đúng theo kỹ năng
        skill_accuracy = {
            skill: round(stats["correct"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
            for skill, stats in skill_stats.items()
        }

        # Tính accuracy cho từng học sinh
        students_list = []
        for sid, data in student_data.items():
            accuracy = round(data["correct_count"] / data["total_questions"] * 100, 1) if data["total_questions"] > 0 else 0
            students_list.append({
                "student_id": sid,
                "experiment_group": data["experiment_group"],
                "total_questions": data["total_questions"],
                "accuracy_pct": accuracy,
                "latest_theta": round(data["latest_theta"], 3)
            })

        return {
            "status": "success",
            "total_sessions": len(logs),
            "total_students": len(students_list),
            "students": students_list,
            "skill_accuracy": skill_accuracy,
            "note": "Dữ liệu từ research_experiment_logs.jsonl — dùng cho báo cáo giáo viên và nghiên cứu KHKT."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi đọc báo cáo giáo viên: {str(e)}")


@app.get("/api/research/export")
async def export_research_data(
    grade: Optional[str] = None,
    experiment_group: Optional[str] = None,
    start_date: Optional[str] = None,   # ISO: YYYY-MM-DD
    end_date: Optional[str] = None,     # ISO: YYYY-MM-DD
    format: Optional[str] = "csv",      # "csv" hoặc "xlsx"
    db: Session = Depends(get_session)
):
    """
    API xuất dữ liệu nghiên cứu thực nghiệm dưới dạng CSV (UTF-8 BOM) hoặc Excel (.xlsx).
    Hỗ trợ lọc theo: khối lớp, nhóm thực nghiệm (ADAPTIVE/CONTROL), khoảng thời gian.
    Dữ liệu thô dùng cho phân tích định lượng SPSS/R trong nghiên cứu KHKT.
    """
    try:
        import csv
        from datetime import datetime, timezone, date as date_type

        # Parse date filters → unix timestamp
        ts_start = None
        ts_end = None
        if start_date:
            try:
                d = date_type.fromisoformat(start_date)
                ts_start = int(datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc).timestamp())
            except ValueError:
                pass
        if end_date:
            try:
                d = date_type.fromisoformat(end_date)
                ts_end = int(datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=timezone.utc).timestamp())
            except ValueError:
                pass

        # 1. Lấy thông tin tất cả user từ DB để map grade và fullname
        users = db.exec(select(User)).all()
        user_map = {u.username.lower(): {"grade": u.grade, "fullname": u.fullname} for u in users}

        # 2. Đọc logs ghi nhận thực tế từ file log thực nghiệm
        logger = ResearchLogger()
        logs = logger.read_all_logs()

        # Các hàng dữ liệu đã lọc
        headers = [
            "student_id", "fullname", "grade", "experiment_group", "repetition_engine",
            "question_id", "skill", "correct", "theta_before", "theta_after",
            "timestamp_unix", "timestamp_iso"
        ]
        rows = []
        for log in logs:
            sid = log.get("student_id", "anonymous").lower()
            grp = log.get("experiment_group", "ADAPTIVE")
            ts = log.get("timestamp", 0)

            u_info = user_map.get(sid, {"grade": "12", "fullname": "Ẩn danh"})
            u_grade = u_info["grade"]
            u_fullname = u_info["fullname"]

            # Áp dụng bộ lọc
            if grade and u_grade != grade:
                continue
            if experiment_group and grp != experiment_group:
                continue
            if ts_start and ts < ts_start:
                continue
            if ts_end and ts > ts_end:
                continue

            dt = datetime.fromtimestamp(ts, tz=timezone.utc)
            iso_str = dt.strftime("%Y-%m-%d %H:%M:%S")
            rows.append([
                sid, u_fullname, u_grade, grp,
                log.get("repetition_engine", "SM2"),
                log.get("question_id", ""),
                log.get("skill", ""),
                log.get("correct", 0),
                log.get("theta_before", 0.0),
                log.get("theta_after", 0.0),
                ts, iso_str
            ])

        now_str = datetime.now().strftime('%Y%m%d_%H%M%S')

        # ── XUẤT EXCEL (.xlsx) ─────────────────────────────────────────
        if format and format.lower() == "xlsx":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Research Data"

                # Style header
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill("solid", fgColor="2B3674")
                thin = Side(style='thin', color="C0C0C0")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.append(headers)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Dữ liệu
                for row in rows:
                    ws.append(row)

                # Auto column width
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                ws.freeze_panes = "A2"

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                filename = f"research_export_{now_str}.xlsx"
                return Response(
                    content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f'attachment; filename="{filename}"',
                        "Cache-Control": "no-cache"
                    }
                )
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl chưa được cài đặt. Chạy: pip install openpyxl")

        # ── XUẤT CSV (mặc định, UTF-8 BOM) ────────────────────────────
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        bom = b'\xef\xbb\xbf'
        response_content = bom + output.getvalue().encode('utf-8')
        filename = f"research_export_{now_str}.csv"
        return Response(
            content=response_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xuất dữ liệu nghiên cứu: {str(e)}")


@app.get("/api/research/theta-timeline")
async def get_theta_timeline(
    grade: Optional[str] = None,
    experiment_group: Optional[str] = None,
    db: Session = Depends(get_session)
):
    """
    API trả về dữ liệu chuỗi thời gian theta trung bình theo ngày.
    Phục vụ biểu đồ tiến trình năng lực trong AdminPanel.
    Trả về: [{date, avg_theta_adaptive, avg_theta_control, count_adaptive, count_control}]
    """
    try:
        from datetime import datetime, timezone

        users = db.exec(select(User)).all()
        user_map = {u.username.lower(): {"grade": u.grade} for u in users}

        logger = ResearchLogger()
        logs = logger.read_all_logs()

        # Nhóm theo ngày và nhóm thực nghiệm
        daily: Dict[str, Dict[str, list]] = defaultdict(lambda: {"ADAPTIVE": [], "CONTROL": []})

        for log in logs:
            sid = log.get("student_id", "anonymous").lower()
            grp = log.get("experiment_group", "ADAPTIVE")
            ts = log.get("timestamp", 0)
            theta_after = float(log.get("theta_after", 0.0))

            u_grade = user_map.get(sid, {}).get("grade", "12")
            if grade and u_grade != grade:
                continue
            if experiment_group and grp != experiment_group:
                continue

            try:
                dt = datetime.fromtimestamp(ts, tz=timezone.utc)
                day_str = dt.strftime("%Y-%m-%d")
            except Exception:
                continue

            if grp in ("ADAPTIVE", "CONTROL"):
                daily[day_str][grp].append(theta_after)

        # Tổng hợp kết quả theo ngày tăng dần
        timeline = []
        for day in sorted(daily.keys()):
            ad = daily[day]["ADAPTIVE"]
            co = daily[day]["CONTROL"]
            timeline.append({
                "date": day,
                "avg_theta_adaptive": round(sum(ad) / len(ad), 4) if ad else None,
                "avg_theta_control": round(sum(co) / len(co), 4) if co else None,
                "count_adaptive": len(ad),
                "count_control": len(co),
            })

        return {
            "status": "success",
            "total_days": len(timeline),
            "timeline": timeline
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi lấy dữ liệu theta-timeline: {str(e)}")


class GenerateSentenceRequest(BaseModel):
    level: str = "A2"

@app.post("/api/pronounce/generate-sentence")
async def generate_pronounce_sentence(request: GenerateSentenceRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    API sinh ngẫu nhiên 1 câu thực hành phát âm Tiếng Anh theo trình độ bằng Gemini AI.
    """
    try:
        data = await generate_pronunciation_sentence_with_gemini(
            level=request.level,
            custom_key=x_gemini_key
        )
        return {
            "status": "success",
            "sentence": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi sinh câu phát âm AI: {str(e)}")




@app.get("/api/questions")
async def get_questions(grade: str):
    """
    API lấy danh sách câu hỏi theo khối lớp từ file questions.json
    """
    try:
        import json
        file_path = os.path.join(os.path.dirname(__file__), "questions.json")
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Không tìm thấy file câu hỏi questions.json")
            
        with open(file_path, "r", encoding="utf-8") as f:
            all_questions = json.load(f)
            
        questions = all_questions.get(grade, [])
        return {
            "status": "success",
            "grade": grade,
            "total": len(questions),
            "questions": questions
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi lấy danh sách câu hỏi: {str(e)}")


class GenerateReadingRequest(BaseModel):
    topic: str
    grade: str = "10"
    theta: float = 0.0

class ScorePredictRequest(BaseModel):
    theta: float
    ef: float
    streak: int
    pronounce_score: float

@app.post("/api/reading/generate")
async def generate_reading_endpoint(request: GenerateReadingRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    API sinh bài đọc tiếng Anh thích ứng theo sở thích và năng lực theta bằng Gemini AI.
    """
    try:
        data = generate_adaptive_reading(
            topic=request.topic,
            grade=request.grade,
            theta=request.theta,
            user_api_key=x_gemini_key
        )
        return {
            "status": "success",
            "reading": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi sinh bài đọc thích ứng: {str(e)}")

@app.post("/api/predict/scores")
async def predict_scores_endpoint(request: ScorePredictRequest):
    """
    API dự đoán điểm số thi chuẩn hóa (THPT Quốc gia, VSTEP, IELTS) bằng thuật toán hồi quy.
    """
    try:
        data = calculate_predicted_exam_scores(
            theta=request.theta,
            ef=request.ef,
            streak=request.streak,
            pronounce_score=request.pronounce_score
        )
        return {
            "status": "success",
            "predictions": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi dự báo điểm số: {str(e)}")

class EvaluateWritingRequest(BaseModel):
    text: str
    prompt: str
    grade: str = "10"

@app.post("/api/writing/evaluate")
async def evaluate_writing_endpoint(request: EvaluateWritingRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    API đánh giá và sửa lỗi chi tiết đoạn văn viết tiếng Anh của học sinh bằng Gemini AI.
    """
    if not request.text.strip():
        raise HTTPException(status_code=400, detail="Văn bản viết không được để trống")
    try:
        data = evaluate_writing_with_gemini(
            student_text=request.text,
            topic_prompt=request.prompt,
            grade=request.grade,
            user_api_key=x_gemini_key
        )
        return {
            "status": "success",
            "evaluation": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi đánh giá bài viết: {str(e)}")


class SampleWritingRequest(BaseModel):
    prompt: str
    grade: str = "10"

@app.post("/api/writing/sample")
async def sample_writing_endpoint(request: SampleWritingRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    API sinh dàn ý gợi ý, từ vựng trọng tâm và bài viết mẫu tham khảo bằng Gemini AI.
    """
    try:
        data = generate_writing_sample_with_gemini(
            topic_prompt=request.prompt,
            grade=request.grade,
            user_api_key=x_gemini_key
        )
        return {
            "status": "success",
            "sample": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi sinh bài mẫu tham khảo: {str(e)}")


class AdaptiveListeningRequest(BaseModel):
    topic: str = "Technology and Environment"
    grade: str = "10"
    theta: float = 0.0

@app.post("/api/generate-adaptive-listening")
async def generate_adaptive_listening_endpoint(request: AdaptiveListeningRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    API sinh bài nghe tiếng Anh thích ứng AI theo chủ đề sở thích và năng lực học sinh.
    """
    try:
        data = generate_adaptive_listening(
            topic=request.topic,
            grade=request.grade,
            theta=request.theta,
            user_api_key=x_gemini_key
        )
        return {
            "status": "success",
            "listening": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi sinh bài nghe thích ứng: {str(e)}")

@app.post("/api/ai/solve-photo")
async def solve_photo_endpoint(
    file: UploadFile = File(...),
    grade: str = Form("12"),
    x_gemini_key: Optional[str] = Header(None)
):
    """
    API Nhận diện câu hỏi tiếng Anh từ ảnh chụp/tải lên và hướng dẫn giải chi tiết từng bước.
    """
    try:
        image_bytes = await file.read()
        mime_type = file.content_type or "image/jpeg"
        result = await solve_exam_by_image(
            image_bytes=image_bytes,
            mime_type=mime_type,
            grade=grade,
            custom_key=x_gemini_key
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi giải đề bằng ảnh: {str(e)}")


class DictLookupRequest(BaseModel):
    word: str
    grade: Optional[str] = "12"

# Bảng tra cứu trực tiếp các từ tiếng Việt thông dụng sang tiếng Anh chuẩn
COMMON_VI_EN_MAP = {
    "con ca": ("fish", "/fɪʃ/", "Danh từ (n.)", "con cá, loài cá sống dưới nước", "Many colorful fish swim in the coral reef.", "Nhiều loài cá sặc sỡ bơi lội trong rạn san hô."),
    "ca": ("fish", "/fɪʃ/", "Danh từ (n.)", "con cá, loài cá sống dưới nước", "Fresh fish is a healthy source of protein.", "Cá tươi là nguồn cung cấp chất đạm lành mạnh."),
    "con meo": ("cat", "/kæt/", "Danh từ (n.)", "con mèo", "The cat is sleeping peacefully under the warm sunshine.", "Con mèo đang ngủ yên bình dưới ánh nắng ấm áp."),
    "meo": ("cat", "/kæt/", "Danh từ (n.)", "con mèo", "Cats are very independent and agile pets.", "Mèo là loài thú cưng rất độc lập và nhanh nhẹn."),
    "con cho": ("dog", "/dɒɡ/", "Danh từ (n.)", "con chó", "Dogs are faithful companions to humans.", "Chó là những người bạn đồng hành trung thành của con người."),
    "cho": ("dog", "/dɒɡ/", "Danh từ (n.)", "con chó", "Dogs are faithful companions to humans.", "Chó là những người bạn đồng hành trung thành của con người."),
    "bac si": ("doctor", "/ˈdɒk.tər/", "Danh từ (n.)", "bác sĩ y khoa", "The doctor examines patients carefully in the hospital.", "Bác sĩ khám bệnh cẩn thận cho bệnh nhân trong bệnh viện."),
    "giao vien": ("teacher", "/ˈtiː.tʃər/", "Danh từ (n.)", "giáo viên, thầy cô giáo", "Our English teacher is enthusiastic and dedicated.", "Giáo viên tiếng Anh của chúng tôi rất nhiệt tình và tận tâm."),
    "thay co": ("teacher", "/ˈtiː.tʃər/", "Danh từ (n.)", "thầy cô giáo", "Students show profound gratitude to their teachers.", "Học sinh bày tỏ lòng biết ơn sâu sắc tới các thầy cô giáo."),
    "hoc sinh": ("student", "/ˈstjuː.dənt/", "Danh từ (n.)", "học sinh, sinh viên", "Diligent students review grammar lessons daily.", "Những học sinh chăm chỉ ôn luyện các bài học ngữ pháp mỗi ngày."),
    "ngoi nha": ("house", "/haʊs/", "Danh từ (n.)", "ngôi nhà", "They built a lovely house with a peaceful garden.", "Họ đã xây một ngôi nhà đáng yêu có khu vườn yên bình."),
    "nha": ("house", "/haʊs/", "Danh từ (n.)", "ngôi nhà", "Welcome to our cozy house.", "Chào mừng bạn đến với ngôi nhà ấm cúng của chúng tôi."),
    "truong hoc": ("school", "/skuːl/", "Danh từ (n.)", "trường học", "Children walk to school together every morning.", "Trẻ em cùng nhau đi bộ đến trường mỗi buổi sáng."),
    "xe dap": ("bicycle", "/ˈbaɪ.sɪ.kəl/", "Danh từ (n.)", "xe đạp", "Riding a bicycle improves cardiovascular health.", "Đi xe đạp giúp cải thiện sức khỏe tim mạch."),
    "xe may": ("motorbike", "/ˈməʊ.tə.baɪk/", "Danh từ (n.)", "xe máy", "Motorbikes are a common means of transport in Vietnam.", "Xe máy là phương tiện giao thông phổ biến tại Việt Nam."),
    "o to": ("car", "/kɑːr/", "Danh từ (n.)", "xe ô tô, xe hơi", "Electric cars reduce urban air pollution.", "Xe ô tô điện giúp làm giảm ô nhiễm không khí đô thị."),
    "xe hoi": ("car", "/kɑːr/", "Danh từ (n.)", "xe ô tô, xe hơi", "Electric cars reduce urban air pollution.", "Xe ô tô điện giúp làm giảm ô nhiễm không khí đô thị."),
    "may tinh": ("computer", "/kəmˈpjuː.tər/", "Danh từ (n.)", "máy tính", "Computers are indispensable tools for digital learning.", "Máy tính là công cụ không thể thiếu cho việc học tập kỹ thuật số."),
    "dien thoai": ("phone", "/fəʊn/", "Danh từ (n.)", "điện thoại di động", "Smartphones connect people around the world.", "Điện thoại thông minh kết nối mọi người trên khắp thế giới."),
    "moi truong": ("environment", "/ɪnˈvaɪ.rən.mənt/", "Danh từ (n.)", "môi trường sinh thái", "Protecting the environment is the responsibility of everyone.", "Bảo vệ môi trường là trách nhiệm của tất cả mọi người."),
    "suc khoe": ("health", "/helθ/", "Danh từ (n.)", "sức khỏe", "Good health is the most valuable asset in life.", "Sức khỏe tốt là tài sản quý giá nhất trong cuộc đời."),
    "gia dinh": ("family", "/ˈfæm.əl.i/", "Danh từ (n.)", "gia đình", "Family is where love begins and never ends.", "Gia đình là nơi tình yêu bắt đầu và không bao giờ kết thúc."),
    "ban be": ("friend", "/frend/", "Danh từ (n.)", "bạn bè, người bạn", "A true friend is always there to support you.", "Một người bạn thực sự luôn ở đó để hỗ trợ bạn."),
    "thanh cong": ("success", "/səkˈses/", "Danh từ (n.)", "sự thành công", "Hard work and perseverance lead to lasting success.", "Chăm chỉ và kiên trì sẽ dẫn tới thành công lâu dài."),
    "hanh phuc": ("happiness", "/ˈhæp.i.nəs/", "Danh từ (n.)", "sự hạnh phúc", "True happiness comes from helping others.", "Hạnh phúc đích thực đến từ việc giúp đỡ người khác."),
    "sang tao": ("creative", "/kriˈeɪ.tɪv/", "Tính từ (adj.)", "sáng tạo, có óc tưởng tượng", "She came up with a creative solution to the problem.", "Cô ấy đã đưa ra một giải pháp sáng tạo cho vấn đề."),
    "thong minh": ("intelligent", "/ɪnˈtel.ɪ.dʒənt/", "Tính từ (adj.)", "thông minh, sáng dạ", "Dolphins are highly intelligent marine creatures.", "Cá heo là loài sinh vật biển vô cùng thông minh."),
    "kien tri": ("perseverance", "/ˌpɜː.sɪˈvɪə.rəns/", "Danh từ (n.)", "sự kiên trì, nhẫn nại", "Through sheer perseverance, he mastered English fluency.", "Nhờ vào sự kiên trì tuyệt đối, anh ấy đã làm chủ được sự trôi chảy trong tiếng Anh."),
    "thoi tiet": ("weather", "/ˈweð.ər/", "Danh từ (n.)", "thời tiết khí hậu", "The weather is pleasant and sunny today.", "Thời tiết hôm nay rất dễ chịu và có nắng ấm."),
    "sach": ("book", "/bʊk/", "Danh từ (n.)", "cuốn sách", "Reading books broadens your knowledge and imagination.", "Đọc sách giúp mở rộng kiến thức và trí tưởng tượng của bạn."),
    "tien": ("money", "/ˈmʌn.i/", "Danh từ (n.)", "tiền bạc", "Money cannot buy genuine happiness.", "Tiền bạc không thể mua được hạnh phúc chân thật."),
    "thoi gian": ("time", "/taɪm/", "Danh từ (n.)", "thời gian", "Time is the most precious resource we possess.", "Thời gian là nguồn tài nguyên quý giá nhất mà chúng ta sở hữu.")
}

def _remove_vietnamese_accents(text: str) -> str:
    import re
    s = text.lower().strip()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    return s

@app.post("/api/dictionary/lookup")
async def dictionary_lookup(request: DictLookupRequest, x_gemini_key: Optional[str] = Header(None)):
    """
    Tra cứu song ngữ thông minh Anh - Việt & Việt - Anh (Smart Bilingual Dictionary Lookup).
    Nhận diện tự động tiếng Việt (VD: "con cá", "bác sĩ", "môi trường") để dịch chuẩn sang tiếng Anh
    kèm đầy đủ phiên âm IPA, loại từ, nghĩa tiếng Việt và câu ví dụ ngữ cảnh có audio.
    """
    raw_query = request.word.strip()
    if not raw_query:
        raise HTTPException(status_code=400, detail="Vui lòng nhập từ cần tra cứu")

    norm_query = _remove_vietnamese_accents(raw_query)

    # 1. Kiểm tra bảng từ điển Tiếng Việt -> Tiếng Anh thông dụng
    if norm_query in COMMON_VI_EN_MAP:
        eng_word, ipa, pos, meaning, ex, ex_vi = COMMON_VI_EN_MAP[norm_query]
        return {
            "status": "success",
            "source": "bilingual_lexicon",
            "data": {
                "word": eng_word,
                "ipa": ipa,
                "pos": pos,
                "meaning": f"{raw_query} → {meaning}",
                "example": ex,
                "example_vi": ex_vi
            }
        }

    # 2. Tra cứu trong SQLite Database (khớp cả từ tiếng Anh lẫn nghĩa tiếng Việt)
    with Session(engine) as db:
        # Khớp tiếng Anh chính xác
        db_word = db.exec(select(VocabularyWord).where(VocabularyWord.word.ilike(raw_query))).first()
        if db_word:
            return {
                "status": "success",
                "source": "database",
                "data": {
                    "word": db_word.word,
                    "ipa": db_word.ipa,
                    "pos": db_word.pos,
                    "meaning": db_word.meaning,
                    "example": db_word.example,
                    "example_vi": db_word.example_vi
                }
            }

        # Khớp nghĩa tiếng Việt trong database
        all_words = db.exec(select(VocabularyWord)).all()
        for w in all_words:
            if norm_query in _remove_vietnamese_accents(w.meaning):
                return {
                    "status": "success",
                    "source": "database_meaning_match",
                    "data": {
                        "word": w.word,
                        "ipa": w.ipa,
                        "pos": w.pos,
                        "meaning": w.meaning,
                        "example": w.example,
                        "example_vi": w.example_vi
                    }
                }

    # 3. Tra cứu bằng Gemini AI (Song ngữ 2 chiều)
    active_key = clean_api_key(x_gemini_key) or clean_api_key(os.getenv("GEMINI_API_KEY"))
    if active_key:
        try:
            genai.configure(api_key=active_key)
            model = genai.GenerativeModel(
                model_name="gemini-1.5-flash",
                generation_config={"response_mime_type": "application/json"}
            )
            prompt = f"""You are an Oxford/Cambridge English-Vietnamese Bilingual Dictionary.
Query from user: "{raw_query}"

Instructions:
1. If the query is in VIETNAMESE (e.g. "con cá", "bác sĩ", "kinh doanh", "môi trường"):
   - Identify the most accurate target ENGLISH word (e.g. "fish", "doctor", "business", "environment").
   - Provide its standard International Phonetic Alphabet (IPA) transcription (e.g. "/fɪʃ/").
   - Provide its part of speech in Vietnamese (e.g. "Danh từ (n.)", "Động từ (v.)", "Tính từ (adj.)").
   - Provide the concise Vietnamese definition of the word.
   - Provide a clear, natural English example sentence.
   - Provide the Vietnamese translation of the example sentence.

2. If the query is in ENGLISH (e.g. "sustainable", "serendipity", "achieve"):
   - Provide standard IPA transcription.
   - Provide part of speech in Vietnamese.
   - Provide the Vietnamese meaning.
   - Provide an English example sentence and its Vietnamese translation.

Return strictly a valid JSON object with keys:
{{
  "word": "<English word>",
  "ipa": "/.../",
  "pos": "<Part of speech in Vietnamese>",
  "meaning": "<Vietnamese meaning>",
  "example": "<English example sentence>",
  "example_vi": "<Vietnamese translation of example>"
}}"""
            res = model.generate_content(prompt)
            data = json.loads(res.text)
            return {"status": "success", "source": "ai", "data": data}
        except Exception as e:
            print(f"Lỗi AI dictionary lookup: {e}")

    # 4. Tra cứu dự phòng trực tuyến qua MyMemory API (Miễn phí 100% không cần key)
    try:
        import urllib.request
        import urllib.parse
        encoded = urllib.parse.quote(raw_query)
        url = f"https://api.mymemory.translated.net/get?q={encoded}&langpair=vi|en"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=4) as response:
            trans_json = json.loads(response.read().decode("utf-8"))
            translated_text = trans_json.get("responseData", {}).get("translatedText", "").strip()
            # Bỏ dấu chấm cuối nếu dịch ra 1 từ
            clean_eng = translated_text.rstrip('.').strip().lower()
            if clean_eng and clean_eng != raw_query.lower():
                return {
                    "status": "success",
                    "source": "online_bilingual_translator",
                    "data": {
                        "word": clean_eng,
                        "ipa": f"/{clean_eng}/",
                        "pos": "Từ vựng tiếng Anh",
                        "meaning": f"{raw_query} (Nghĩa tiếng Anh: {clean_eng})",
                        "example": f"Students should learn how to use '{clean_eng}' in daily communication.",
                        "example_vi": f"Học sinh nên rèn luyện cách sử dụng từ '{clean_eng}' ({raw_query}) trong giao tiếp hàng ngày."
                    }
                }
    except Exception as e:
        print("Lỗi MyMemory fallback:", e)

    # 5. Fallback cuối cùng
    return {
        "status": "success",
        "source": "offline_dict",
        "data": {
            "word": raw_query,
            "ipa": f"/{raw_query.lower()}/",
            "pos": "Từ vựng tra cứu",
            "meaning": f"Nghĩa tra cứu của từ '{raw_query}'",
            "example": f"Practice using '{raw_query}' regularly to expand your vocabulary.",
            "example_vi": f"Hãy luyện tập sử dụng từ '{raw_query}' thường xuyên để mở rộng vốn từ vựng."
        }
    }


if __name__ == "__main__":
    import uvicorn
    # Chạy server cổng 8000 lắng nghe trên toàn bộ interface mạng cục bộ
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
